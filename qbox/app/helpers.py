import csv
from flask import send_file, Response, flash
def read_csv():    
    with open('qbox/app/xlsx/master.csv', mode='r') as infile:
        infile = csv.reader(infile)
        for line in infile:
            box, archivio, scaffale, progetto, dipartimento, codice, volume_da, volume_a = line[0].split(';')
            print(box, progetto, codice, volume_a)


#import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def export_db():
    try:
        records = db.session.query(Volume).all() # filter only active, exclude x - not found site
        wb = load_workbook('app/GTF-GPS/GTF-GPS-COR-24034-01 Archival Records Storage Form_English.xlsx')
        ws = wb.active
        row_count = 0
        for el in records:
            row_count +=1
            new_row = [
                
                el.box.name, 
                el.box.section.area.site.city,
                el.box.section.area.site.country,
                el.group.name,
                el.type.name,
                el.type.retention_code,
                el.type.description,
                el.date_start,
                el.date_end,
                el.project.code,
                el.project.name,
                el.name,
                el.type.activation_on, 
                el.activation_date,
                el.type.retention_days,
                el.endlife_date,
                el.type.security_class,
                el.account_number,
                #el.order_number,
                el.box.section.area.site.name,
                #el.box.section.area.name,
                #el.request_by,
                #el.request_on,
                #el.box.section.name
                ]
            ws.append(new_row)
            
        wb.save('app/xlsx/export_db.xlsx')    
        #b = BytesIO(open(outfile,'rb'))
            #outfile.seek(0)
            #outfile.close()
            
        return send_file('xlsx/export_db.xlsx', as_attachment=True, download_name='GTF-GPS-COR-24034-01 Archival Records Storage Form_English.xlsx')
    except Exception as e:
        flash('254 Export Error:', e)
        return False

def exportexcel(query):
    try:
        records = query
        if len(records) > 15000:
            flash("Limit of 15K reached: " + str(len(records)),'info')
            return False
        wb = load_workbook('app/xlsx/Export_2.xlsx')
        ws = wb.active
        row_count = 0
        for el in records:
            row_count +=1
            new_row = [
                str(el.box.id),
                el.box.name, 
                el.box.section.area.site.city,
                el.box.section.area.site.country,
                el.group.name,
                el.type.name,
                el.type.retention_code,
                el.type.description,
                el.date_start,
                el.date_end,
                el.project.code,
                el.project.name,
                el.name,
                el.type.activation_on, 
                el.activation_date,
                el.type.retention_days,
                el.endlife_date,
                el.type.security_class,
                el.account_number,
                el.order_number,
                el.box.section.area.site.name,
                el.box.section.area.name,
                el.request_by,
                el.request_on,
                el.box.section.name]
            ws.append(new_row)
        
        
        tab = Table(displayName="Table1", ref="A2:AA"+str(row_count+2))
        labels = ['Doc ID',	
                'Box N.',	
                'Città',	
                'Nazione',	
                'Gruppo / Dipartimento',	
                'Codice archivio',	
                'Periodo di conservazione', 
                'Descrizione del codice archivio',	
                'Data da',	
                'Data a',	
                'N. progetto',	
                'Nome del progetto',	
                'Descrizione',	
                'Riferimento di attivazione',	
                'Data di attivazione',
                'GG  conservazione',
                'Data distruzione', 
                'Livello di sicurezza',	
                'Numero di conto',	
                'Commessa N.',	
                'Archivio',	
                'Stanza',	
                'Richiedente',
                'data ritiro',	
                'Scaffale N.', 
                'Persona',
                'Data']
        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(name="TableStyleMedium9", 
                                #showFirstColumn=False,
                                #showLastColumn=False, 
                                
                                #showColumnStripes=True,
                                showRowStripes=True)
        tab.tableStyleInfo = style 

        '''
        Table must be added using ws.add_table() method to avoid duplicate names.
        Using this method ensures table name is unque through out defined names and all other table name. 
        '''
        #tab.headerRowCount = False  
        tab._initialise_columns()
        for column, value in zip(tab.tableColumns, labels):
            column.name = value

        ws.add_table(tab)
        wb.save('app/xlsx/export_search.xlsx')    
        #b = BytesIO(open(outfile,'rb'))
            #outfile.seek(0)
            #outfile.close()
            
        return send_file('xlsx/export_search.xlsx', as_attachment=True, download_name='ADM Search Export.xlsx')
    except Exception as e:
        flash('254 Export Error:', e)
        return False


def write_csv(query):
    with open('app/static/downloads/result.csv','w') as outfile: 
        outcsv = csv.writer(outfile, dialect='excel', lineterminator="\n", quotechar='"')
        if len(query[1]) > 20000:    
            flash('Error: Too many Records selected: '+ str(len(query[1])),'info' )
            return False
        #print(query)
        labels = ['NADA_ID','BOX_CODE','CITY','COUNTRY','GROUP','TYPE',
                      'TYPE_CODE','TYPE_DESCRIPTION','DATE START','DATE_END',
                      'PROJECT','PROJECT NAME','DESCRIPTION','TRIGGER ON', 'TRIGGER DATE',
                      'RETENTION DAYS','ENDLIFE DATE','SECURITY CLASS','ACCOUNT','ORDER',
                      'SITE','AREA','REQUEST BY','REQUEST ON', 'SECTION']
        outcsv.writerow(labels)
        for el in query[1]:
            print(type(el), el)
            #input('look here')
            
            outcsv.writerow([
                el.box.id,
                el.box.name, 
                el.box.section.area.site.city,
                el.box.section.area.site.country,
                el.group.name,
                el.type.name,
                el.type.retention_code,
                el.type.description,
                el.date_start,
                el.date_end,
                el.project.code,
                el.project.name,
                el.name,
                el.type.activation_on, 
                el.activation_date,
                el.type.retention_days,
                el.endlife_date,
                el.type.security_class,
                el.account_number,
                el.order_number,
                el.box.section.area.site.name,
                el.box.section.area.name,
                el.request_by,
                el.request_on,
                el.box.section.name
                  
                  ])
        #b = BytesIO(open(outfile,'rb'))
        #outfile.seek(0)
        #outfile.close()
           
    return send_file('static/downloads/result.csv', as_attachment=True, download_name='result.csv')
    

from random import choice
def random_color():
    return 'rgb('+ str(choice(range(0,255))) + ', ' + str(choice(range(0,255)))  + ', ' + str(choice(range(0,255))) + ', 0.5)'
        
from openpyxl import load_workbook
from .models import Account, Project, Site, Area, Section, Group, Box, Volume, Type, Move
from app import db


from datetime import datetime
def to_date(date):
    if date:
        return datetime.strptime(date, "%d/%m/%Y")
    return None
    


def load_master():
    wb = load_workbook('app/xlsx/master.xlsx')
    ws = wb.active
    count = 0
    

    for row in ws.iter_rows():
        count += 1
        box_id = row[0].value
        city = row[1].value
        country = row[2].value
        archivio = row[1].value
        scaffale = row[2].value
        progetto = row[3].value
        dipartimento = row[4].value
        codice = row[5].value
        volume_da = row[6].value
        volume_a = row[7].value
        
        account = db.session.query(Account).filter(Account.name == 'TPIT').first()
        if not account:
            account = Account(name = 'TPIT', created_by_fk = 1, changed_by_fk = 1)
        db.session.add(account)
        
        project = db.session.query(Project).filter(Project.name == progetto).first()
        if not project:
            project = Project(name = progetto, account= account, created_by_fk = 1, changed_by_fk = 1)
        db.session.add(project)
        
        site = db.session.query(Site).filter(Site.name == archivio).first()
        if not site:
            site = Site(name = archivio, country='Italy', city='Rome', address='Address', account= account, created_by_fk = 1, changed_by_fk = 1)
        db.session.add(site)
            
        area = db.session.query(Area).filter(Area.name == site.name).first()
        if not area:
            area = Area(name = site.name, site= site, created_by_fk = 1, changed_by_fk = 1)
        db.session.add(area)    
        
        section = db.session.query(Section).filter(Section.name == scaffale).first()
        if not section:
            section = Section(name = scaffale, area= area, created_by_fk = 1, changed_by_fk = 1)
        db.session.add(section)
        
        box = db.session.query(Box).filter(Box.name == box_id).first()
        if not box:
            box = Box(name = box_id, section= section, created_by_fk = 1, changed_by_fk = 1)
        db.session.add(box)
            
        group = db.session.query(Group).filter(Group.name == dipartimento).first()
        if not group:
            group = Group(name = dipartimento, account= account, created_by_fk = 1, changed_by_fk = 1)
        db.session.add(group)
        
        type = db.session.query(Type).filter(Type.name == codice).first()
        if not type:
            type = Type(name = codice, account= account, created_by_fk = 1, changed_by_fk = 1)
        db.session.add(type)
        
        volume = db.session.query(Volume).filter(
            Volume.v_from == volume_da,
            Volume.v_to == volume_a,
            Volume.box == box
            ).first()
        if not volume:
            volume = Volume(v_from = volume_da,
                            v_to = volume_a,
                            project = project,
                            box = box,
                            group = group,
                            type = type,
                            name = str(str(volume_da) + " " + str(volume_a)),
                            created_by_fk = 1, 
                            changed_by_fk = 1
                            )
        print(count, str(volume.name))
        db.session.add(volume)
        
    
        db.session.commit()
        

def load_master_2():
    wb = load_workbook('app/xlsx/master_2.xlsx', data_only=True)
    ws = wb.active
    count = 0
    errors = []
    

    for row in ws.iter_rows(min_row=3):
        try:
            count += 1
            box_id = row[0].value
            city = row[1].value
            country = row[2].value
            dipartimento = row[3].value or 'ToBeDefined'
            codice = row[4].value or 'ToBeDefined'
            periodo_cons = row[5].value
            descrizione_code = row[6].value or 'ToBeDefined'
            date_start = row[7].value or None
            date_end = row[8].value or None
            progetto = row[9].value or 'ToBeDefined'
            nome_progetto = row[10].value or 'ToBeDefined'
            description = row[11].value or 'ToBeDefined'
            rif_attivazione = row[12].value or None
            data_attivazione = row[13].value  or None# date_end + 1 gg
            gg_conservazione = row[14].value or None
            data_distruzione = row[15].value or None# data_attivazione + gg_conservazione
            livello_sicurezza = row[16].value or 'ToBeDefined'
            
            numero_conto = row[17].value or None
            commessa = None
            
            archivio = row[19].value or "ND"
            stanza = row[20].value or "ND"
            
            richiedente = row[21].value or "ND"
            data_ritiro = row[22].value or None
            
            scaffale = row[23].value or "ND"
            
            move_to = None
            move_date = None
            
            
            
            
            
            #volume_a = row[7].value
            
            account = db.session.query(Account).filter(Account.name == 'TPIT').first()
            if not account:
                account = Account(name = 'TPIT', created_by_fk = 1, changed_by_fk = 1)
            db.session.add(account)
            
            project = db.session.query(Project).filter(Project.code == progetto).first()
            if not project:
                project = Project(name = nome_progetto, code = progetto, account= account, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(project)
            
            site = db.session.query(Site).filter(Site.name == archivio).first()
            if not site:
                site = Site(name = archivio, country=country, city=city, address='ND', account= account, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(site)
                
            area = db.session.query(Area).filter(Area.name == stanza).first()
            if not area:
                area = Area(name = stanza, site= site, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(area)    
            
            section = db.session.query(Section).filter(Section.name == scaffale).first()
            if not section:
                section = Section(name = scaffale, area= area, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(section)
            
            box = db.session.query(Box).filter(Box.name == box_id).first()
            if not box:
                box = Box(name = box_id, section= section, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(box)
                
            group = db.session.query(Group).filter(Group.name == dipartimento).first()
            if not group:
                group = Group(name = dipartimento, account= account, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(group)
            
            type = db.session.query(Type).filter(Type.name == codice).first()
            if not type:
                type = Type(name = codice, 
                            account= account,
                            retention_code = periodo_cons,
                            description = descrizione_code,
                            activation_on = rif_attivazione,
                            retention_days = gg_conservazione,
                            security_class = livello_sicurezza,
                            created_by_fk = 1, 
                            changed_by_fk = 1)
            db.session.add(type)
            
            volume = db.session.query(Volume).filter(
                Volume.name == description,
                Volume.box == box
                ).first()
            if not volume:
                volume = Volume(name = description,
                                project = project,
                                box = box,
                                group = group,
                                type = type,
                                date_start =date_start,
                                date_end = date_end,
                                activation_date = data_attivazione,
                                endlife_date = data_distruzione,
                                account_number = numero_conto,
                                order_number = commessa,
                                request_by = richiedente,
                                request_on = data_ritiro,
                                created_by_fk = 1, 
                                changed_by_fk = 1
                                )
            print(count, str(volume.name))
            db.session.add(volume)
            
            
            print(project.color)
            db.session.flush()
            print(project.color)
            db.session.commit()
        except Exception as e:
            #row[30] = str(e)
            db.session.rollback() 
            print(e,'         *-*-*-*-*-*-*-*-  ERROR *-*-*-*-*-*-')
            #errors.append((count+2,e))
            #pass
            break
           
    print('')
    print('         *-*-*-*-*-*-*-*-  LIST ERRORS *-*-*-*-*-*-')               
    for e in errors:
        print(e)    
                   
'''
2500	2554		31/12/00
2001	2499		31/12/08
1920	2000		31/12/00
1760	1919		31/12/00
1436	1759		31/12/90
1001	1435		31/12/80
'''
def activation_date(prj_code):
    dates = [
        [2500,2554,datetime(2000,12,31)],
        [2001,2499,datetime(2008,12,31)],
        [1760,2000,datetime(2000,12,31)],
        [1436,1759,datetime(1990,12,31)],
        [1401,1435,datetime(1990,12,31)]
    ]
    for date in dates:
        try:
            print('project code:', prj_code, date)
            if int(prj_code) >= date[0] and int(prj_code) <= date[1]:
                #print('Match !')
                return date[2]

        except Exception as e:
            print(e)
            return datetime(2024,2,28)
    #print('NOT MATCH', type(prj_code), type(date[0]), type([1]))
    return datetime(2024,2,28)


def load_master_3A():
    wb = load_workbook('app/xlsx/master_6.xlsx', data_only=True)
    ws = wb.active
    count = 0
    errors = []
    

    for row in ws.iter_rows(min_row=3):
        try:
            #s_date = activation_date(row[11].value or 10)
            
            count += 1
            box_id = row[0].value
            city = row[1].value
            country = row[2].value
            dipartimento = row[3].value or 'ToBeDefined'
            codice = row[4].value or 'ToBeDefined'
            periodo_cons = row[5].value
            descrizione_code = row[6].value or 'ToBeDefined'
            date_start = row[7].value 
            date_end = row[8].value 
            progetto = row[9].value or 'ToBeDefined'
            nome_progetto = row[10].value or 'ToBeDefined'
            description = row[11].value or 'ToBeDefined'
            rif_attivazione = row[12].value or None
            data_attivazione = row[13].value # date_end + 1 gg
            gg_conservazione = row[14].value or None
            data_distruzione = row[15].value # data_attivazione + gg_conservazione
            livello_sicurezza = row[16].value or 'ToBeDefined'
            
            numero_conto = row[17].value or None
            commessa = row[18].value or None
            
            #add calculated column to the table
            archivio = row[19].value or "ND"
            
            stanza = row[20].value or "ND"
            
            richiedente = row[21].value or None
            data_ritiro = row[22].value or None
            
            scaffale = row[23].value or "ND"
            
            move_to = row[25].value or None
            move_date = row[26].value or None
            
            
            
            
            
            #volume_a = row[7].value
            
            account = db.session.query(Account).filter(Account.name == 'TPIT').first()
            if not account:
                account = Account(name = 'TPIT', created_by_fk = 1, changed_by_fk = 1)
            db.session.add(account)
            
            project = db.session.query(Project).filter(Project.code == progetto).first()
            if not project:
                project = Project(name = nome_progetto, code = progetto, account= account, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(project)
            
            site = db.session.query(Site).filter(Site.name == archivio).first()
            if not site:
                site = Site(name = archivio, country=country, city=city, address='ND', account= account, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(site)
                
            area = db.session.query(Area).filter(Area.name == stanza).first()
            if not area:
                area = Area(name = stanza, site= site, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(area)    
            
            section = db.session.query(Section).join(Area).filter(Section.name == scaffale,
                                                       Area.name == stanza).first()
            if not section:
                section = Section(name = scaffale, area= area, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(section)
            
            box = db.session.query(Box).filter(Box.name == box_id).first()
            if not box:
                box = Box(name = box_id, section= section, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(box)
                
            group = db.session.query(Group).filter(Group.name == dipartimento).first()
            if not group:
                group = Group(name = dipartimento, account= account, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(group)
            
            type = db.session.query(Type).filter(Type.name == codice).first()
            if not type:
                type = Type(name = codice, 
                            account= account,
                            retention_code = periodo_cons,
                            description = descrizione_code,
                            activation_on = rif_attivazione,
                            retention_days = gg_conservazione,
                            security_class = livello_sicurezza,
                            created_by_fk = 1, 
                            changed_by_fk = 1)
            db.session.add(type)
            
            volume = db.session.query(Volume).filter(
                Volume.name == description,
                Volume.box == box
                ).first()
            if not volume:
                volume = Volume(name = description,
                                project = project,
                                box = box,
                                group = group,
                                type = type,
                                date_start =date_start,
                                date_end = date_end,
                                activation_date = data_attivazione,
                                endlife_date = data_distruzione,
                                account_number = numero_conto,
                                order_number = commessa,
                                request_by = richiedente,
                                request_on = data_ritiro,
                                created_by_fk = 1, 
                                changed_by_fk = 1
                                )
            print(count, str(volume.name))
            db.session.add(volume)
            
            
        
            db.session.commit()
        except Exception as e:
            #row[30] = str(e)
            db.session.rollback() 
            print(e,'         *-*-*-*-*-*-*-*-  ERROR *-*-*-*-*-*-')
            errors.append((count+2,e))
            input('look')
            #pass
            break
           
    print('')
    print('         *-*-*-*-*-*-*-*-  LIST ERRORS *-*-*-*-*-*-')               
    for e in errors:
        print(e)    

def load_master_3():
    wb = load_workbook('app/xlsx/LAST .xlsx', data_only=True)
    ws = wb.active
    count = 0
    errors = []
    

    for row in ws.iter_rows(min_row=3):
        try:
            count += 1
            box_id = row[0].value
            city = "Rome"
            country = "Italy"
            dipartimento = row[4].value or 'ToBeDefined'
            codice = row[6].value or 'ToBeDefined'
            periodo_cons = 'ToBeDefined'
            descrizione_code = row[5].value or 'ToBeDefined'
            date_start = None
            date_end = None
            progetto = row[3].value or 'ToBeDefined'
            nome_progetto = 'ToBeDefined'
            description = row[8].value or "ND " + ' Vol: ' + row[9].value or "ND " + " " + row[10].value or 'ToBeDefined'
            rif_attivazione = 'ToBeDefined'
            data_attivazione = None # date_end + 1 gg
            gg_conservazione = 0
            data_distruzione = None # data_attivazione + gg_conservazione
            livello_sicurezza = 'ToBeDefined'
            
            numero_conto = None
            commessa = 'ToBeDefined'
            
            archivio = row[11].value or 'ToBeDefined'
            stanza = row[2].value or 'ToBeDefined'
            
            richiedente = 'ToBeDefined'
            data_ritiro = None
            
            scaffale = row[1].value or 'ToBeDefined'
            
            move_to = 'ToBeDefined'
            move_date = None
            
            
            
            
            
            #volume_a = row[7].value
            
            account = db.session.query(Account).filter(Account.name == 'TPIT').first()
            if not account:
                account = Account(name = 'TPIT', created_by_fk = 1, changed_by_fk = 1)
            db.session.add(account)
            
            project = db.session.query(Project).filter(Project.code == progetto).first()
            if not project:
                project = Project(name = nome_progetto, code = progetto, account= account, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(project)
            
            site = db.session.query(Site).filter(Site.name == archivio).first()
            if not site:
                site = Site(name = archivio, country=country, city=city, address='ND', account= account, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(site)
                
            area = db.session.query(Area).filter(Area.name == stanza).first()
            if not area:
                area = Area(name = stanza, site= site, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(area)    
            
            section = db.session.query(Section).filter(Section.name == scaffale).first()
            if not section:
                section = Section(name = scaffale, area= area, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(section)
            
            box = db.session.query(Box).filter(Box.name == box_id).first()
            if not box:
                box = Box(name = box_id, section= section, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(box)
                
            group = db.session.query(Group).filter(Group.name == dipartimento).first()
            if not group:
                group = Group(name = dipartimento, account= account, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(group)
            
            type = db.session.query(Type).filter(Type.name == codice).first()
            if not type:
                type = Type(name = codice, 
                            account= account,
                            retention_code = periodo_cons,
                            description = descrizione_code,
                            activation_on = rif_attivazione,
                            retention_days = gg_conservazione,
                            security_class = livello_sicurezza,
                            created_by_fk = 1, 
                            changed_by_fk = 1)
            db.session.add(type)
            
            volume = db.session.query(Volume).filter(
                Volume.name == description,
                Volume.box == box
                ).first()
            if not volume:
                volume = Volume(name = description,
                                project = project,
                                box = box,
                                group = group,
                                type = type,
                                date_start =date_start,
                                date_end = date_end,
                                activation_date = data_attivazione,
                                endlife_date = data_distruzione,
                                account_number = numero_conto,
                                order_number = commessa,
                                request_by = richiedente,
                                request_on = data_ritiro,
                                created_by_fk = 1, 
                                changed_by_fk = 1
                                )
            print(count, str(volume.name))
            db.session.add(volume)
            
            
        
            db.session.commit()
        except Exception as e:
            row[13].value = str(e)
            db.session.rollback() 
            print('         *-*-*-*-*-*-*-*-  ERROR *-*-*-*-*-*-', count, e)
            errors.append((count+2,e))
            pass
    
    wb.save('LAST .xlsx')
    print('')
    print('         *-*-*-*-*-*-*-*-  LIST ERRORS *-*-*-*-*-*-')               
    for e in errors:
        print(e)    
                   
 
def load_master_30K():
    wb = load_workbook('app/xlsx/30K_Da normalizzare 1.xlsx', data_only=True)
    ws = wb.active
    count = 0
    errors = []
    print('LOAD MASTER 30K - 1')

    for row in ws.iter_rows(min_row=3):
        try:
            s_date = activation_date(row[3].value or 10)
            print('LOAD MASTER 30K - 2')
            count += 1
            box_id = row[0].value
            city = "Rome"
            country = "Italy"
            dipartimento = row[4].value or 'ToBeDefined'
            codice = row[11].value or 'ToBeDefined'
            periodo_cons = row[12].value or 'ToBeDefined'
            descrizione_code = row[13].value or 'ToBeDefined'
            date_start = s_date
            date_end = s_date
            progetto = row[3].value or 'ToBeDefined'
            nome_progetto = 'ToBeDefined'
            description = row[8].value or "ND " + ' Vol: ' + row[9].value or "ND " + " " + row[10].value or 'ToBeDefined'
            rif_attivazione = 'ToBeDefined'
            data_attivazione = s_date
            gg_conservazione = 0
            data_distruzione = s_date
            livello_sicurezza = 'ToBeDefined'
            
            numero_conto = None
            commessa = 'ToBeDefined'
            
            archivio = row[26].value or 'ToBeDefined'
            stanza = row[1].value or 'ToBeDefined'
            
            richiedente = 'ToBeDefined'
            data_ritiro = None
            
            scaffale = row[2].value or 'ToBeDefined'
            
            move_to = 'ToBeDefined'
            move_date = None
            
            
            
            
            
            #volume_a = row[7].value
            
            account = db.session.query(Account).filter(Account.name == 'TPIT').first()
            if not account:
                account = Account(name = 'TPIT', created_by_fk = 1, changed_by_fk = 1)
            db.session.add(account)
            
            project = db.session.query(Project).filter(Project.code == progetto).first()
            if not project:
                project = Project(name = nome_progetto, code = progetto, account= account, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(project)
            
            site = db.session.query(Site).filter(Site.name == archivio).first()
            if not site:
                site = Site(name = archivio, country=country, city=city, address='ND', account= account, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(site)
                
            area = db.session.query(Area).filter(Area.name == stanza).first()
            if not area:
                area = Area(name = stanza, site= site, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(area)    
            
            section = db.session.query(Section).join(Area).filter(Section.name == scaffale,
                                                       Area.name == stanza).first()
            
            if not section:
                print(box_id,'New SECTION')
                section = Section(name = scaffale, area= area, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(section)
            
            print('BOX ADD - 1 section', section.name, section.area)
            box = db.session.query(Box).filter(Box.name == box_id).first()
            if not box:
                print(box_id,'New BOX')
                box = Box(name = box_id, section= section, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(box)
            print(box_id,'BOX ADDED', box.name, section.name)    
            group = db.session.query(Group).filter(Group.name == dipartimento).first()
            if not group:
                group = Group(name = dipartimento, account= account, created_by_fk = 1, changed_by_fk = 1)
            db.session.add(group)
            
            type = db.session.query(Type).filter(Type.name == codice).first()
            if not type:
                type = Type(name = codice, 
                            account= account,
                            retention_code = periodo_cons,
                            description = descrizione_code,
                            activation_on = rif_attivazione,
                            retention_days = gg_conservazione,
                            security_class = livello_sicurezza,
                            created_by_fk = 1, 
                            changed_by_fk = 1)
            db.session.add(type)
            
            volume = db.session.query(Volume).filter(
                Volume.name == description,
                Volume.box == box
                ).first()
            if not volume:
                volume = Volume(name = description,
                                project = project,
                                box = box,
                                group = group,
                                type = type,
                                date_start =date_start,
                                date_end = date_end,
                                activation_date = data_attivazione,
                                endlife_date = data_distruzione,
                                account_number = numero_conto,
                                order_number = commessa,
                                request_by = richiedente,
                                request_on = data_ritiro,
                                created_by_fk = 1, 
                                changed_by_fk = 1
                                )
            print(count, str(volume.name))
            db.session.add(volume)
            
            
        
            db.session.commit()
        except Exception as e:
            print('LOAD MASTER 30K - 3')
            db.session.rollback()
            row[13].value = str(e)
            db.session.rollback() 
            print('         *-*-*-*-*-*-*-*-  ERROR *-*-*-*-*-*-', count, e)
            errors.append((count+2,e))
            pass
    
    wb.save('30K_Da normalizzare 1.xlsx')
    print('')
    print('         *-*-*-*-*-*-*-*-  LIST ERRORS *-*-*-*-*-*-')               
    for e in errors:
        print(e)    
                   


def project_mass_color():
    projects = db.session.query(Project).all()
    for project in projects:
        project.changed_by_fk = '1'
        project.color = random_color()
    
    db.session.commit()

from datetime import timedelta
def update_endlife():            
    # update endlife date
    print('**** * * * * * * update endlife date')
    volumes = db.session.query(Volume).all()
    #print('len volumes:', len(volumes))
    count = 0
    if volumes:
        for v in volumes:
            count += 1
            days = v.type.retention_days
            #print('hold date:', v.activation_date)
            #v.activation_date = activation_date(v.project.code)
            v.endlife_date = v.activation_date + timedelta(days=days)
            v.changed_by_fk = '1'
            print(count,v.id, v.activation_date, v.endlife_date)
            db.session.add(v)
    try:
        db.session.commit()
        
    except Exception as e:
            print(e)
            db.session.rollback()
                
            

def update_date2():
    volumes = db.session.query(Volume).filter(
        Volume.endlife_date == None
    ).all()
    print('**********************************')
    print('len volumes:', len(volumes))
    print('')
    print('')
    count = 0
    for v in volumes:
        count += 1
        days = v.type.retention_days
        v.endlife_date = v.activation_date + timedelta(days=days)
        v.changed_by_fk = '1'
        print(count,v.id, v.activation_date, v.endlife_date)
        
    db.session.commit()
     
        
            
def active_box():
    tot = 26467
    count = 0
    not_found = []
    with open('app/xlsx/act_box.csv') as file:
        for code in file:
            count += 1
            code = code.strip()
            
            print('Code:',code, type(code), len(code),str(count) +'/' + str(tot))
            box = db.session.query(Box).filter(Box.name == code, Box.active == False).first()
            if box:
                box.active = True
                box.changed_by_fk = '1'
                try:
                    db.session.commit()
                except Exception as e:
                    print(e)
                    db.session.rollback()
            else:
                print(code, 'Not Found')
                not_found.append([code])
    print('NOT FOUND LIST')
    print(not_found)       


def notfound_box():
    tot = 26467
    count = 0
    not_found = []
    with open('app/xlsx/act_box.csv') as file:
        for code in file:
            count += 1
            code = code.strip()
            
            print('Code:',code,str(count) +'/' + str(tot))
            box = db.session.query(Box).filter(Box.name == code).first()
            if box:
                print(box, 'found')
                if box.active == False:
                    print(' ++++++++++ + + + + + + + +++++ BOX INACTIVE FOUND')
                    box.active = True
                    box.changed_by_fk = '1'
                    db.session.commit()
                
            else:
                print(code, 'Not Found')
                not_found.append([code])
          
    with open('nada_notfound2.csv','w') as f:
        writer = csv.writer(f)
        writer.writerows(not_found)           
        



def load_notfound():
    count = 0
    errors = []
    
    with open('nada_notfound2.csv') as csv:
        for code in csv:
            count += 1
            code = code.strip()
            
            with open('app/xlsx/archivio29-04-2024.csv') as csv:
                for line in csv:
                    line_lst = line.split(';')
                    
                    if line_lst[0] == code:
                        try:
                            if line_lst[3]:
                            # è un volume
                            
                                s_date = activation_date(line_lst[3])
                                
                                count += 1
                                box_id = code
                                city = 'Rome'
                                country = 'Italy'
                                dipartimento = line_lst[4] or 'ToBeDefined'
                                codice = 'ToBeDefined'
                                periodo_cons = 'ToBeDefined'
                                descrizione_code = 'ToBeDefined'
                                date_start = s_date
                                date_end = s_date
                                progetto = line_lst[3]
                                nome_progetto = 'ToBeDefined'
                                description = " - ".join(line_lst[3:])
                                rif_attivazione = None
                                data_attivazione = s_date# date_end + 1 gg
                                gg_conservazione = None
                                data_distruzione = s_date# data_attivazione + gg_conservazione
                                livello_sicurezza = 'ToBeDefined'
                                
                                numero_conto =  None
                                commessa = None
                                
                                archivio = line_lst[1]
                                
                                richiedente = None
                                data_ritiro = None
                                move_to = None
                                move_date = None
                            
                                stanza = 'ToBeDefined'
                                scaffale = 'ToBeDefined'
                            
                    
                                account = db.session.query(Account).filter(Account.name == 'TPIT').first()
                                if not account:
                                    account = Account(name = 'TPIT', created_by_fk = 1, changed_by_fk = 1)
                                db.session.add(account)
                                
                                project = db.session.query(Project).filter(Project.code == progetto).first()
                                if not project:
                                    project = Project(name = nome_progetto, code = progetto, account= account, created_by_fk = 1, changed_by_fk = 1)
                                db.session.add(project)
                                
                                site = db.session.query(Site).filter(Site.name == archivio).first()
                                if not site:
                                    site = Site(name = archivio, country=country, city=city, address='ND', account= account, created_by_fk = 1, changed_by_fk = 1)
                                db.session.add(site)
                                    
                                area = db.session.query(Area).filter(Area.name == stanza).first()
                                if not area:
                                    area = Area(name = stanza, site= site, created_by_fk = 1, changed_by_fk = 1)
                                db.session.add(area)    
                                
                                section = db.session.query(Section).join(Area).filter(Section.name == scaffale,
                                                                        Area.name == stanza).first()
                                if not section:
                                    section = Section(name = scaffale, area= area, created_by_fk = 1, changed_by_fk = 1)
                                db.session.add(section)
                                
                                box = db.session.query(Box).filter(Box.name == box_id).first()
                                if not box:
                                    box = Box(name = box_id, 
                                              section= section,
                                              active = True, 
                                              created_by_fk = 1, 
                                              changed_by_fk = 1)
                                db.session.add(box)
                                    
                                group = db.session.query(Group).filter(Group.name == dipartimento).first()
                                if not group:
                                    group = Group(name = dipartimento, account= account, created_by_fk = 1, changed_by_fk = 1)
                                db.session.add(group)
                                
                                type = db.session.query(Type).filter(Type.name == codice).first()
                                if not type:
                                    type = Type(name = codice, 
                                                account= account,
                                                retention_code = periodo_cons,
                                                description = descrizione_code,
                                                activation_on = rif_attivazione,
                                                retention_days = gg_conservazione,
                                                security_class = livello_sicurezza,
                                                created_by_fk = 1, 
                                                changed_by_fk = 1)
                                db.session.add(type)
                                
                                volume = db.session.query(Volume).filter(
                                    Volume.name == description,
                                    Volume.box == box
                                    ).first()
                                if not volume:
                                    volume = Volume(name = description,
                                                    project = project,
                                                    box = box,
                                                    group = group,
                                                    type = type,
                                                    date_start =date_start,
                                                    date_end = date_end,
                                                    activation_date = data_attivazione,
                                                    endlife_date = data_distruzione,
                                                    account_number = numero_conto,
                                                    order_number = commessa,
                                                    request_by = richiedente,
                                                    request_on = data_ritiro,
                                                    created_by_fk = 1, 
                                                    changed_by_fk = 1
                                                    )
                                print(count, str(volume.name))
                                db.session.add(volume)
                                
                            try:    
                            
                                db.session.commit()
                            except Exception as e:
                                #row[30] = str(e)
                                db.session.rollback() 
                                print(e,'         *-*-*-*-*-*-*-*-  ERROR *-*-*-*-*-*-')
                                errors.append((count+2,code, e))
                                #pass
                                break
                        
                        except Exception as e:
                            print('EXCEPTION PROJECT')
                            pass
                
            print('')
            print('         *-*-*-*-*-*-*-*-  LIST ERRORS *-*-*-*-*-*-')               
            for e in errors:
                print(e) 
                
                
def update_position():
    
    count = 0
    errors = [] 
            
    with open('app/xlsx/box_position.csv') as csv:
        for line in csv:
            count += 1
            try:
                box_code, box_site, box_area, box_section = line.split(';')
                print(count, box_code, box_site, box_area, box_section)
                
                box = db.session.query(Box).filter(Box.name == box_code).first()
                if box:
                    site = db.session.query(Site).filter(Site.name == box_site).first()
                    if not site:
                        account = db.session.query(Account).filter(Account.name == 'TPIT').first()
                        site = Site(name = box_site, country='Italy', city='Rome', address='ND', account= account, created_by_fk = 1, changed_by_fk = 1)
                        db.session.add(site)
                            
                    area = db.session.query(Area).join(Site).filter(Area.name == box_area, Site.name == box_site).first()
                    if not area:
                        area = Area(name = box_area, site= site, created_by_fk = 1, changed_by_fk = 1)
                        db.session.add(area)    
                        
                    section = db.session.query(Section).join(Area, Site).filter(
                                                    Section.name == box_section,
                                                    Site.name == box_site,
                                                    Area.name == box_area).first()
                    if not section:
                        section = Section(name = box_section, area= area, created_by_fk = 1, changed_by_fk = 1)
                        db.session.add(section)
                        #db.session.flush()
                    hold_section = box.section
                    
                    #area.site = site
                    #section.area = area
                    box.section = section
                    
                    #area.changed_by_fk = '1'
                    #section.changed_by_fk = '1'
                    box.changed_by_fk = '1'
                    print('| SECTION CHANGE:', box.id,box.name,'-- from: ',hold_section ,'to: ',box.section)

                    db.session.add(box)
                    db.session.commit()
                
                else:
                    errors.append([box_code, 'NOT FOUND'])       
                    
            except Exception as e:
                print(e)
                errors.append(e)
                i = input('Something Happen... enter to continue or b to break -> ')
                if i == 'b':
                    break
                
    
    print('*-*-* -* -* -* - * -* - * -*  ERRORS -*-* * - * - -- * * - -* * - * ')
    for error in errors:
        print(error)      
            
import os
         
def db_to_csv():
        
    # MySQL connection information
    mysql_host = 'your_host'
    mysql_user = 'root'
    mysql_password = 'lollipop300777'
    mysql_database = 'qbox5'

    # Define your SQL query
    sql_query = '''
    select
	volume.id,
    box.name as Box_Code,
    section.name as Scaffale,
    area.name as Archivio,
    
    site.city,
    site.country,
    group.name,
    type.name as Codice_Archivio,
    type.retention_code as Conservazione,
    type.description as Descrizione_Codice,
	volume.date_start as Data_Da,
    volume.date_end as Data_A,
    project.code as Project_Code,
    project.name as Project_Name,
   
    volume.name as Descrizione,
    
    type.activation_on as Rif_Attivazione,
    volume.activation_date as Activation_Date,
    type.retention_days as Retention_Days,
    volume.endlife_date as Data_Distruzione,
    type.security_class as Livello_Sicurezza,
    volume.account_number as Numero_Conto
    
    

from qbox5.volume
join qbox5.box on volume.box_id = box.id
join qbox5.section on box.section_id = section.id
join qbox5.area on section.area_id = area.id
join qbox5.site on area.site_id = site.id
join qbox5.group on volume.group_id = group.id
join qbox5.type on volume.type_id = type.id
join qbox5.project on volume.project_id = project.id
    
'''
    
    '''
    # query for sal
    
    * aggiungere lo user del commento e gli user che hanno lavorato i task
    * prevedere la possibilità di avere questo formato per user
    * fare il trim del nome del deliverable
    * considerare inclusione pending
    *  
    
    select isa.projecttask.name as activity, isa.deliverable.name as deliverable, sum(isa.task.duration) as duration, group_concat(isa.comments.text) as note
    from isa.task

    join isa.deliverable on isa.task.deliverable_id = isa.deliverable.id
    join isa.projecttask on isa.deliverable.projecttask_id = isa.projecttask.id
    left join isa.comments on isa.task.id = isa.comments.task_id

    WHERE startDate BETWEEN '2013-03-12 00:00:00' AND '2013-03-12 23:59:59'
    AND endDate BETWEEN '2013-03-12 00:00:00' AND '2013-03-12 23:59:59'
  
    group by isa.projecttask.name, isa.deliverable.name
    '''

    # Define the output CSV file path
    output_csv_file = 'app/static/downloads/output.csv' 

    # Use the mysqldump command to execute the query and save the result as a CSV file
    #command = f"mysql -h {mysql_host} -u {mysql_user} -p{mysql_password} {mysql_database} -e \"{sql_query}\" > {output_csv_file}"
    command = f"mysql -u {mysql_user} -p{mysql_password} {mysql_database} -e '{sql_query}' > {output_csv_file}"
    
    # Execute the command
    
    os.system(command)
    
   

    print(f"Query result has been saved to {output_csv_file}")
    return send_file('static/downloads/output.csv', as_attachment=True, download_name='result.csv')



def update_desc():
    
    count = 0
    errors = [] 
            
    with open('app/xlsx/update_desc.csv') as csv:
        for line in csv:
            count += 1
            try:
                vol_id, box_code, vol_name = line.split(';')
                print(count, vol_id, box_code, vol_name)
                
                v = db.session.query(Volume).get(vol_id)
                if v:
                    print('Hold:',v.name,'New:',vol_name)
                    v.name = vol_name
                    v.changed_by_fk = '1'
                    
                    db.session.add(v)
                    db.session.commit()
                
                else:
                    errors.append([box_code, 'NOT FOUND'])       
                    
            except Exception as e:
                print(e)
                errors.append(e)
                i = input('Something Happen... enter to continue or b to break -> ')
                if i == 'b':
                    break
                
    
    print('*-*-* -* -* -* - * -* - * -*  ERRORS -*-* * - * - -- * * - -* * - * ')
    for error in errors:
        print(error)  
        
        
def load_volumes(): 
    count = 0
    errors = []
    boxes = db.session.query(Box).all()
    if boxes:
        for box in boxes:
            count += 1
            code = box.name
            
            with open('app/xlsx/archivio29-04-2024.csv') as csv:
                for line in csv:
                    line_lst = line.split(';')
                    #print('db_Code',code)
                    #print('a_code',line_lst[0])
                    #print('line',line_lst)
                    #input('look here')
                    if line_lst[0] == code:
                        #print('Code Match')
                        try:
                            if line_lst[2] == '':
                                #print('Is a Volume')
                            # è un volume

                                s_date = activation_date(line_lst[3])
                                
                                count += 1
                                #box_id = code
                                city = 'Rome'
                                country = 'Italy'
                                dipartimento = line_lst[4] or 'ToBeDefined'
                                codice = 'ToBeDefined'
                                periodo_cons = 'ToBeDefined'
                                descrizione_code = 'ToBeDefined'
                                date_start = s_date
                                date_end = s_date
                                progetto = line_lst[3] or 'ToBeDefined'
                                nome_progetto = 'ToBeDefined'
                                description = " - ".join(line_lst[3:])
                                rif_attivazione = None
                                data_attivazione = s_date# date_end + 1 gg
                                gg_conservazione = None
                                data_distruzione = s_date# data_attivazione + gg_conservazione
                                livello_sicurezza = 'ToBeDefined'
                                
                                numero_conto =  None
                                commessa = None
                                
                                archivio = line_lst[1]
                                
                                richiedente = None
                                data_ritiro = None
                                move_to = None
                                move_date = None
                            
                                stanza = 'ToBeDefined'
                                scaffale = 'ToBeDefined'
                            
                    
                                account = db.session.query(Account).filter(Account.name == 'TPIT').first()
                                if not account:
                                    account = Account(name = 'TPIT', created_by_fk = 1, changed_by_fk = 1)
                                db.session.add(account)
                                
                                project = db.session.query(Project).filter(Project.code == progetto).first()
                                if not project:
                                    project = Project(name = nome_progetto, code = progetto, account= account, created_by_fk = 1, changed_by_fk = 1)
                                db.session.add(project)
                                
                                '''
                                site = db.session.query(Site).filter(Site.name == archivio).first()
                                if not site:
                                    site = Site(name = archivio, country=country, city=city, address='ND', account= account, created_by_fk = 1, changed_by_fk = 1)
                                db.session.add(site)
                                    
                                area = db.session.query(Area).filter(Area.name == stanza).first()
                                if not area:
                                    area = Area(name = stanza, site= site, created_by_fk = 1, changed_by_fk = 1)
                                db.session.add(area)    
                                
                                section = db.session.query(Section).join(Area).filter(Section.name == scaffale,
                                                                        Area.name == stanza).first()
                                if not section:
                                    section = Section(name = scaffale, area= area, created_by_fk = 1, changed_by_fk = 1)
                                db.session.add(section)
                                
                                box = db.session.query(Box).filter(Box.name == box_id).first()
                                if not box:
                                    box = Box(name = box_id, 
                                              section= section,
                                              active = True, 
                                              created_by_fk = 1, 
                                              changed_by_fk = 1)
                                db.session.add(box)
                                '''
                                group = db.session.query(Group).filter(Group.name == dipartimento).first()
                                if not group:
                                    group = Group(name = dipartimento, account= account, created_by_fk = 1, changed_by_fk = 1)
                                db.session.add(group)
                                
                                type = db.session.query(Type).filter(Type.name == codice).first()
                                if not type:
                                    type = Type(name = codice, 
                                                account= account,
                                                retention_code = periodo_cons,
                                                description = descrizione_code,
                                                activation_on = rif_attivazione,
                                                retention_days = gg_conservazione,
                                                security_class = livello_sicurezza,
                                                created_by_fk = 1, 
                                                changed_by_fk = 1)
                                db.session.add(type)
                                
                                volume = Volume(name = description,
                                                project = project,
                                                box = box,
                                                group = group,
                                                type = type,
                                                date_start =date_start,
                                                date_end = date_end,
                                                activation_date = data_attivazione,
                                                endlife_date = data_distruzione,
                                                account_number = numero_conto,
                                                order_number = commessa,
                                                request_by = richiedente,
                                                request_on = data_ritiro,
                                                created_by_fk = 1, 
                                                changed_by_fk = 1
                                                )
                                print(count, str(volume.name))
                                db.session.add(volume)
                            
                        except Exception as e:
                            print('EXCEPTION PROJECT',e)
                            errors.append(e)
                            pass
                                    
        try:    
        
            db.session.commit()
        except Exception as e:
            #row[30] = str(e)
            db.session.rollback() 
                                
            print(e,'         *-*-*-*-*-*-*-*-  ERROR *-*-*-*-*-*-')
            errors.append((count+2,code, e))
            #pass
            
                        
                        
                
    print('')
    print('         *-*-*-*-*-*-*-*-  LIST ERRORS *-*-*-*-*-*-')               
    for e in errors:
        print(e) 
        
from flask import url_for, redirect  
def update_request_by():
    volumes = db.session.query(Volume).all()
    # 2115 - INGEGNERIA - SEGRETERIA TECNICA - PETRANGELI - MR/OFFERTA - 0110.01 - 1020.01 
    for v in volumes:
        try:
            print('-> ',v.name)
            req =  v.name.split(' - ',6)[3]
            print(req)
            v.request_by = req
            v.changed_by_fk = '1'
        except Exception as e:
            print(e)
            input('Look... -<')
    db.session.commit()
    
# Chart JS Helper    
# Funzione per creare la vista di elenco dei dati filtrati

def chart_to_csv(title,param):
    #filtered_records = db.session.query(Site).filter(Site.name.contains(param)).all()
    print('-*--* -/* - *-TITLE , PARAM +--++ ',title,param)
    if title == 'site':
        filtered_records = db.session.query(Volume).join(Box,Section,Area,Site).filter(
            Site.name == param).all()
    elif title == 'group':
        filtered_records = db.session.query(Volume).join(Group).filter(
            Group.name == param,
            Site.name != 'X - Not Found').all()
    elif title == 'year':
        filtered_records = db.session.query(Volume).filter(
            Volume.endlife_date.year == param,
            Site.name != 'X - Not Found').all()
        print('Volume for this year:',title, len(filtered_records))

    return exportexcel(filtered_records)
    '''
    with open('app/static/downloads/{}_result.csv'.format(title),'w') as outfile: 
        outcsv = csv.writer(outfile)
        if len(filtered_records) > 3000:    
            flash('Error: Too many Records selected: '+ str(len(filtered_records)),'info' )
            return False
        
        if filtered_records:
            labels = ['NADA_ID','BOX_CODE','CITY','COUNTRY','GROUP','TYPE',
                      'TYPE_CODE','TYPE_DESCRIPTION','DATE START','DATE_END',
                      'PROJECT','PROJECT NAME','DESCRIPTION','TRIGGER ON', 'TRIGGER DATE',
                      'RETENTION DAYS','ENDLIFE DATE','SECURITY CLASS','ACCOUNT','ORDER',
                      'SITE','AREA','REQUEST BY','REQUEST ON', 'SECTION']
            outcsv.writerow(labels)
            for el in filtered_records:
                #print(type(el), el)
                #input('look here')
                
                outcsv.writerow([el.box.id, el.box.name,
                    el.box.section.area.site.city,
                    el.box.section.area.site.country,
                    el.group.name,
                    el.type.name,
                    el.type.retention_code,
                    el.type.description,
                    el.date_start,
                    el.date_end,
                    el.project.code,
                    el.project.name,
                    el.name,
                    el.type.activation_on, 
                    el.activation_date,
                    el.type.retention_days,
                    el.endlife_date,
                    el.type.security_class,
                    el.account_number,
                    el.order_number,
                    el.box.section.area.site.name,
                    el.box.section.area.name,
                    el.request_by,
                    el.request_on,
                    el.box.section.name
                    
                    ])
            
            return send_file('static/downloads/{}_result.csv'.format(title), as_attachment=True, download_name='{}_result.csv'.format(title))
        
        return False
        '''


def import_new_documents():
    
    wb = load_workbook('app/xlsx/new_box2024.xlsx', data_only=True)
    ws = wb.active
    count_match = 0
    count_new = 0
    count_tot = 0
    found = []
    

    for row in ws.iter_rows(min_row=3):
        print(row[0].value)
        count_tot += 1
        
        box = db.session.query(Box).filter(Box.name == row[0].value).first()
        
        if box:
            found.append(box)
            count_match += 1 
            
            
        if box is None:
            count_new += 1
            
    print(found)
    print('Tot:', count_tot,'Count Match',count_match, 'New:', count_new)
    
    
def strip_box_code():
    boxes = db.session.query(Project).all()   
    print(len(boxes))
    count = 0
    for box in boxes:
        count += 1
        
        new_name = box.code.strip()
        box.code = new_name
        box.changed_by_fk = '1'
        
        print(count,'/',len(boxes))
    db.session.commit()


def project_detail():
    wb = load_workbook('app/xlsx/project_list3.xlsx')
    ws = wb.active
    count = 0
    count_found = 0
    for row in ws.iter_rows(min_row=2):
        if row[2].value is not None:
            count += 1
            obj = ''
            
            labels = ['No.',
                    'Contract Entity',
                    'Project Number',	
                    'Project Manager',
                    'OEC',
                    'Client name',
                    'Project name',
                    'Country',
                    'Site',
                    'Contract Value - MMUSD',
                    'Project Category',
                    'SoW',
                    'Association Type',
                    'Award Date',
                    'Completion Date (planned)',
                    'Dif. Award-> Completion (Months)',
                    'Status']
            for l in labels:
                
                #print(l,row[labels.index(l)].value)
                obj += l+': '+ str(row[labels.index(l)].value) + '\n'
            
            print(obj)
            project_number = str(row[2].value).strip()
            project_name = row[6].value
            project_date = row[14].value
            project_date_from = row[13].value
            
            prj = db.session.query(Project).filter(Project.code == project_number).first()
            if prj:
                count_found += 1
                print('Project Found', project_number)
                prj.name = project_name
                prj.note = obj
                prj.changed_by_fk = '1'
                
                db.session.commit()
                prj_docs = db.session.query(Volume).filter(Volume.project_id == prj.id).all()
                
                for doc in prj_docs:
                    doc.date_start = project_date_from
                    doc.date_end = project_date
                    doc.activation_date = project_date
                    doc.changed_by_fk = '1'
                    
                
            else:
                print('Project Not Found:', project_number)
    
    db.session.commit()
    print(' -+-+-+-+-+ RESULT +-+-+-+-')
    print('Count:',count,'Found:',count_found)
    
    
def to_destroy_export(query):
    wb = load_workbook('app/xlsx/GTF-GPS-COR-24036-01-B_2.xlsx')
    ws = wb.active
    count = 0
    count_found = 0
    docs = query
    print(len(docs))
    client_name = ''
    project_title = ''
    system = ''
    volume = ''
    path = ''
    for doc in docs:
        client_name = client_from_prj_note(doc.project.id)
        new_row = [doc.project.account.name, 
                   doc.project.code, 
                   client_name, 
                   doc.project.name, 
                   doc.group.name,
                   system,
                   volume,
                   path,
                   doc.box.name,
                   doc.type.description,
                   doc.type.name,
                   doc.type.retention_code,
                   doc.date_start,
                   doc.date_end,
                   doc.activation_date
                   ]
        print(new_row)
        ws.append(new_row)
    wb.save('app/xlsx/ADM_GTF-GPS-COR-24036-01b.xlsx')
    return send_file('xlsx/ADM_GTF-GPS-COR-24036-01b.xlsx', as_attachment=True, download_name='GTF-GPS-COR-24036-01 Records Destruction Form.xlsx')


def client_from_prj_note(prj_id):
    prj = db.session.query(Project).get(prj_id)
    try:
        for line in prj.note.splitlines():
            label, value = line.split(':',1)
            if label == 'Client name':
                return value.strip() 
    except Exception as e:
        print(e)
        return ''